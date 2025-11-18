import pandas as pd
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import streamlit as st
from io import BytesIO

st.set_page_config(page_title="Accidents File Comparator", layout="centered")
st.title("PT Files Comparator - Accidents")
st.write("Upload the SAP File and the Power BI file.")

selection = st.radio(
    "Choose your filter type:",
    ["PT", "Contracts & Private Hire", "Schools"],
    index = 0
)

        

st.markdown("<h2 style='text-align:center;color:#27AE60;'>Upload SAP File</h2>", unsafe_allow_html=True)
st.write("Upload the SAP file as originally exported. In case the filter type changed, reupload the file.")
uploaded_file1 = st.file_uploader("File 1 SAP File", type=["xlsx", "xls", "xlsm", "xlsb"], key="file1")
st.markdown("<h2 style='text-align:center;color:#2E86C1;'>Upload Power BI File</h2>", unsafe_allow_html=True)
st.write("Upload the Power BI file as originally exported.")
uploaded_file2 = st.file_uploader("File 2 Power BI File", type=["xlsx", "xls", "xlsm", "xlsb"], key="file2")

if uploaded_file1 and uploaded_file2:
    try:
        file1 = pd.read_excel(uploaded_file1, usecols=["Notification", "Equipment", "Accident Date", "Responsible Operations"], dtype=str)
        file1_name = uploaded_file1.name
        if selection == "PT":
            file1 = file1[
                (file1["Responsible Operations"].str.strip().str.lower().isin(["psv", "metro"]))
            ]
        elif selection == "Contracts & Private Hire":
            file1 = file1[
                (file1["Responsible Operations"].str.strip().str.lower().isin(["private hire", "contracts"]))
            ]
        elif selection == "Schools":
            file1 = file1[
                (file1["Responsible Operations"].str.strip().str.lower().isin(["sec"]))
            ]
        file2 = pd.read_excel(uploaded_file2, usecols=["Notification Number", "Fleet No.", "Accident Date"], dtype=str)
        file2_name = uploaded_file2.name
        file2['Notification Number'] = file2['Notification Number'].str.lstrip('0')
        file1['Accident Date'] = pd.to_datetime(file1['Accident Date'], errors='coerce')
        file1['Accident Date'] = file1['Accident Date'].dt.strftime('%d/%m/%Y')
        file1['Accident Date'] = file1['Accident Date'].fillna('')
        file2['Accident Date'] = pd.to_datetime(file2['Accident Date'], errors='coerce')
        file2['Accident Date'] = file2['Accident Date'].dt.strftime('%d/%m/%Y')
        file2['Accident Date'] = file2['Accident Date'].fillna('')



        merged = pd.merge(file1, file2, left_on="Notification", right_on="Notification Number", how="outer", suffixes=('_file1', '_file2'))

        MissingErrors = 0
        EquipmentErrors = 0
        DateErrors = 0
        error_IDs = set()
        analytics_data = []

        for _, row in merged.iterrows():
            num = row["Notification"]

            if pd.isna(num):
                continue

            # Missing in file1
            if pd.isna(row["Equipment"]) and pd.isna(row["Accident Date_file1"]):
                analytics_data.append({
                    "Notification": num,
                    "Responsible Operations": row["Responsible Operations"],
                    "Type": "Missing Notification Number",
                    f"{file1_name}": "Missing",
                    f"{file2_name}": ""
                })
                MissingErrors += 1
                error_IDs.add(num)

            # Missing in file2
            elif pd.isna(row["Fleet No."]) and pd.isna(row["Accident Date_file2"]):
                analytics_data.append({
                    "Notification": num,
                    "Responsible Operations": row["Responsible Operations"],
                    "Type": "Missing Notification Number",
                    f"{file1_name}": "",
                    f"{file2_name}": "Missing"
                })
                MissingErrors += 1
                error_IDs.add(num)

            # Equipment mismatch
            elif row["Equipment"] != row["Fleet No."]:
                analytics_data.append({
                    "Notification": num,
                    "Responsible Operations": row["Responsible Operations"],
                    "Type": "Equipment Mismatch",
                    f"{file1_name}": row["Equipment"],
                    f"{file2_name}": row["Fleet No."]
                })
                EquipmentErrors += 1
                error_IDs.add(num)

            # Date mismatch
            elif row["Accident Date_file1"] != row["Accident Date_file2"]:
                analytics_data.append({
                    "Notification": num,
                    "Responsible Operations": row["Responsible Operations"],
                    "Type": "Date Mismatch",
                    f"{file1_name}": row["Accident Date_file1"],
                    f"{file2_name}": row["Accident Date_file2"]
                })
                DateErrors += 1
                error_IDs.add(num)

        valid_numbers = merged["Notification"].dropna().unique()
        total_IDs = len(valid_numbers)
        total_error_IDs = len(error_IDs)

        if total_error_IDs > 0:
            percent_error = (total_error_IDs / total_IDs) * 100
            accuracy = ((total_IDs - total_error_IDs) / total_IDs) * 100
        else:
            percent_error = 0
            accuracy = 100

        summary_table = [
            ["Total accidents", total_IDs, ""],
            ["Accuracy", f"{accuracy:.2f}%", ""],
            ["Accidents with mismatches", f"{percent_error:.2f}%", ""],
            ["Missing Notification Number case(s)", MissingErrors, ""],
            ["Equipment mismatch case(s)", EquipmentErrors, ""],
            ["Date mismatch case(s)", DateErrors, ""]
        ]

        df_comparison = pd.DataFrame(analytics_data)

        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            sheet_name = "Accidents Comparison"
            df_comparison.to_excel(writer, sheet_name=sheet_name, index=False, startrow=len(summary_table) + 3)
            worksheet = writer.sheets[sheet_name]

            worksheet.merge_cells('A1:D1')
            cell = worksheet['A1']
            cell.value = "Comparison Result"
            cell.font = Font(size=14, bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

            for i, row in enumerate(summary_table, start=3):
                for j, value in enumerate(row, start=1):
                    cell = worksheet.cell(row=i, column=j, value=value)
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            for i, col in enumerate(worksheet.columns, start=1):
                max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
                col_letter = get_column_letter(i)
                worksheet.column_dimensions[col_letter].width = max_length + 4

            start_row = len(summary_table) + 3
            for row in worksheet.iter_rows(
                min_row = start_row,
                max_row = start_row + len(df_comparison) + 1,
                min_col = 1,
                max_col = len(df_comparison.columns)
            ):
                for cell in row:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
        st.markdown("<h4 style='color:#FF0000;'>Enter file name for download</h4>", unsafe_allow_html=True)
        filename = st.text_input("File Name", value=f"{selection}_Accidents_Comparison.xlsx")

        st.download_button(
            label = "💾 Download Comparison Excel",
            data = output.getvalue(),
            file_name = filename,
            mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        st.error(f"⚠️ Error: {str(e)}")